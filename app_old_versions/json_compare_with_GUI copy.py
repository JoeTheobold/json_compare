import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import json

def compare_json(json1, json2, path=""):
    differences = []
    if isinstance(json1, dict) and isinstance(json2, dict):
        all_keys = set(json1.keys()).union(set(json2.keys()))
        for key in all_keys:
            new_path = f"{path}/{key}" if path else key
            if key in json1 and key in json2:
                differences.extend(compare_json(json1[key], json2[key], new_path))
            elif key in json1:
                differences.append((new_path, json1[key], None))
            else:
                differences.append((new_path, None, json2[key]))
    elif isinstance(json1, list) and isinstance(json2, list):
        for i, (item1, item2) in enumerate(zip(json1, json2)):
            differences.extend(compare_json(item1, item2, f"{path}[{i}]"))
        if len(json1) > len(json2):
            for i in range(len(json2), len(json1)):
                differences.append((f"{path}[{i}]", json1[i], None))
        elif len(json2) > len(json1):
            for i in range(len(json1), len(json2)):
                differences.append((f"{path}[{i}]", None, json2[i]))
    else:
        if json1 != json2:
            differences.append((path, json1, json2))
    return differences

def generate_diff():
    json1_file = filedialog.askopenfilename(title="Select First JSON File", filetypes=[("JSON Files", "*.json")])
    json2_file = filedialog.askopenfilename(title="Select Second JSON File", filetypes=[("JSON Files", "*.json")])
    if not json1_file or not json2_file:
        return

    output_file = filedialog.asksaveasfilename(title="Save Excel File As", defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if not output_file:
        return

    with open(json1_file, 'r') as f1, open(json2_file, 'r') as f2:
        json1 = json.load(f1)
        json2 = json.load(f2)

    differences = compare_json(json1, json2)
    df = pd.DataFrame(differences, columns=["Path", "JSON1", "JSON2"])

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Differences')
        workbook = writer.book
        worksheet = writer.sheets['Differences']

        # Apply header font
        header_font = Font(bold=True)
        for col in range(1, 4):
            worksheet.cell(row=1, column=col).font = header_font

        # Highlight differences
        diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for row in range(2, len(df) + 2):
            for col in range(2, 4):
                if worksheet.cell(row=row, column=col).value is not None:
                    worksheet.cell(row=row, column=col).fill = diff_fill

        # Set text wrapping, alignment, and column widths and borders
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
         
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left', vertical='center')

        worksheet.column_dimensions['A'].width = 45  # Path column width
        worksheet.column_dimensions['B'].width = 90  # JSON1 column width
        worksheet.column_dimensions['C'].width = 90  # JSON2 column width

    messagebox.showinfo("Success", f"Differences have been exported to {output_file}")


# Set up the main application window
root = tk.Tk()
root.title("JSON Difference Tool")

tk.Label(root, text="Compare two JSON files and export the differences to an Excel file").pack(pady=10)
tk.Button(root, text="Select Files and Generate Differences", command=generate_diff).pack(pady=20)

root.mainloop()
