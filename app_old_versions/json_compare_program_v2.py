import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import json

def load_json(file_path):
    with open(file_path, 'r') as file:
        return json.load(file)
    
if __name__ == "__main__":
    # Replace 'file1.json' and 'file2.json' with your JSON file paths
    json1 = load_json('nvidia_srv_102.json')
    json2 = load_json('nvidia_srv_103.json')


# Function to recursively compare JSON objects and return differences
def compare_json(json1, json2, path=""):
    differences = []
    if isinstance(json1, dict) and isinstance(json2, dict):
        all_keys = set(json1.keys()).union(set(json2.keys()))
        for key in all_keys:
            new_path = f"{path}/{key}" if path else key
            if key in json1 and key in json2:
                differences.extend(compare_json(json1[key], json2[key], new_path))
            elif key in json1:
                differences.append((new_path, json1[key], None))
            else:
                differences.append((new_path, None, json2[key]))
    elif isinstance(json1, list) and isinstance(json2, list):
        for i, (item1, item2) in enumerate(zip(json1, json2)):
            differences.extend(compare_json(item1, item2, f"{path}[{i}]"))
        if len(json1) > len(json2):
            for i in range(len(json2), len(json1)):
                differences.append((f"{path}[{i}]", json1[i], None))
        elif len(json2) > len(json1):
            for i in range(len(json1), len(json2)):
                differences.append((f"{path}[{i}]", None, json2[i]))
    else:
        if json1 != json2:
            differences.append((path, json1, json2))
    return differences

# Find differences
differences = compare_json(json1, json2)

# Create DataFrame
df = pd.DataFrame(differences, columns=["Path", "JSON1", "JSON2"])

# Export to Excel
output_file = "json_differences.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Differences')
    workbook = writer.book
    worksheet = writer.sheets['Differences']

    # Format the header
    header_font = Font(bold=True)
    for col in range(1, 4):
        worksheet.cell(row=1, column=col).font = header_font

    # Apply conditional formatting
    diff_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in range(2, len(df) + 2):
        for col in range(2, 4):
            if worksheet.cell(row=row, column=col).value is not None:
                worksheet.cell(row=row, column=col).fill = diff_fill

print(f"Differences have been exported to {output_file}")
